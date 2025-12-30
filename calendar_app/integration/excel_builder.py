"""
Excel集成 - 使用openpyxl构建和填充Excel工作簿
"""

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

from calendar_app.config.calendar_config import CalendarConfig
from calendar_app.models.calendar_models import YearCalendarData, CellInfo
from calendar_app.services.cell_image_service import CellImageService
from calendar_app.services.file_manager import FileManager
from calendar_app.models.calendar_models import ImageGenerationRequest


class ExcelBuilder:
    """Excel工作簿构建器"""
    
    def __init__(self, config: CalendarConfig = CalendarConfig):
        self.config = config
        self.file_manager = FileManager(config)
        self.image_service = CellImageService(config)
        self.workbook = None
        self.worksheet = None
    
    def create_workbook(self) -> Workbook:
        """
        创建工作簿
        
        Returns:
            Workbook: openpyxl工作簿对象
        """
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "年历"
        return self.workbook
    
    def setup_layout(self):
        """设置工作簿的布局和页面设置"""
        if not self.worksheet:
            raise ValueError("工作簿未初始化")
        
        # 设置页面
        self.worksheet.page_setup.paperSize = self.worksheet.PAPERSIZE_A3
        self.worksheet.page_setup.orientation = 'landscape'
        self.worksheet.print_options.horizontalCentered = True
        self.worksheet.sheet_view.showGridLines = False
        
        # 设置页边距
        self.worksheet.page_margins.left = self.config.MARGIN_LEFT
        self.worksheet.page_margins.right = self.config.MARGIN_RIGHT
        self.worksheet.page_margins.top = self.config.MARGIN_TOP
        self.worksheet.page_margins.bottom = self.config.MARGIN_BOTTOM
        
        # 设置列宽
        date_column_width = self.config.get_date_column_width()
        for col_num in range(1, self.config.DAYS_PER_MONTH_MAX + 1):
            col_letter = get_column_letter(col_num)
            self.worksheet.column_dimensions[col_letter].width = date_column_width
        
        # 设置行高
        row_height = self.config.get_row_height_points()
        spacer_height = row_height * self.config.MONTH_SPACER_HEIGHT_RATIO
        for month in range(1, 13):
            row = month + (month - 1)
            self.worksheet.row_dimensions[row].height = row_height
            if month < 12:
                self.worksheet.row_dimensions[row + 1].height = spacer_height
    
    def fill_cells(self, calendar_data: YearCalendarData):
        """
        填充日历数据到工作簿
        
        Args:
            calendar_data: 日历数据对象
        """
        if not self.worksheet:
            raise ValueError("工作簿未初始化")
        
        # 创建临时目录
        self.file_manager.create_temp_dir()
        
        # 定义样式
        weekend_fill = PatternFill(
            start_color=self.config.COLOR_WEEKEND_BG, 
            end_color=self.config.COLOR_WEEKEND_BG, 
            fill_type='solid'
        )
        weekday_fill = PatternFill(
            start_color=self.config.COLOR_WEEKDAY_BG, 
            end_color=self.config.COLOR_WEEKDAY_BG, 
            fill_type='solid'
        )
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        no_border = Border()
        
        # 遍历所有月份
        for month_data in calendar_data.months:
            # 遍历该月所有格子
            for cell_info in month_data.cells:
                row = cell_info.row
                col = cell_info.col
                
                # 获取Excel单元格
                cell = self.worksheet.cell(row=row, column=col)
                
                # 日期格子
                cell.fill = weekend_fill if cell_info.is_weekend else weekday_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # 生成格子图像
                self._generate_and_insert_cell_image(cell, cell_info)

        # 清理月份间隔行的边框（无网格线）
        for month in range(1, 12):
            spacer_row = month * 2
            for col in range(1, self.config.DAYS_PER_MONTH_MAX + 1):
                spacer_cell = self.worksheet.cell(row=spacer_row, column=col)
                spacer_cell.border = no_border
                spacer_cell.fill = PatternFill(fill_type=None)
    
    def _generate_and_insert_cell_image(self, cell, cell_info: CellInfo):
        """
        为单个格子生成图像并插入到Excel
        
        Args:
            cell: Excel单元格
            cell_info: 格子信息
        """
        # 创建图像生成请求
        request = ImageGenerationRequest(
            month=cell_info.month,
            day=cell_info.day,
            weekday_char=cell_info.weekday_char,
            cell_width_px=cell_info.width_px,
            cell_height_px=cell_info.height_px,
            is_weekend=cell_info.is_weekend
        )
        
        # 生成图像
        img = self.image_service.create_image(request)
        
        # 保存图像
        img_path = self.file_manager.get_temp_image_path(cell_info.month, cell_info.day)
        img.save(img_path)
        
        # 插入到Excel
        try:
            xl_img = XLImage(img_path)
            start = AnchorMarker(
                col=cell_info.col - 1,
                colOff=0,
                row=cell_info.row - 1,
                rowOff=0,
            )
            end = AnchorMarker(
                col=cell_info.col,
                colOff=0,
                row=cell_info.row,
                rowOff=0,
            )
            xl_img.anchor = TwoCellAnchor(_from=start, to=end, editAs="oneCell")
            self.worksheet.add_image(xl_img)
        except Exception as e:
            print(f"插入图像失败 ({cell_info.month}月{cell_info.day}日): {e}")
    
    def save(self, filename: str) -> bool:
        """
        保存工作簿
        
        Args:
            filename: 输出文件名
            
        Returns:
            bool: 是否成功保存
        """
        try:
            if not self.workbook:
                raise ValueError("工作簿未初始化")
            
            self.workbook.save(filename)
            return True
        except Exception as e:
            print(f"保存文件失败: {e}")
            return False
